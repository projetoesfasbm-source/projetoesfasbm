# backend/services/xlsx_service.py

from __future__ import annotations
from io import BytesIO
from typing import Any, Iterable, Optional
from datetime import date
import locale

# Importações de openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils.cell import range_boundaries

# Configuração de locale para garantir o nome do mês em português
try:
    locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
except locale.Error:
    try:
        locale.setlocale(locale.LC_TIME, 'Portuguese_Brazil') # Windows
    except locale.Error:
        print("Aviso: Não foi possível definir o locale para pt_BR.")

__all__ = ["gerar_mapa_gratificacao_xlsx", "gerar_quadro_horario_xlsx"]

# --- Helpers ---
def _safe(obj: Any, path: str, default: Any = None) -> Any:
    """Navega de forma segura em dicionários ou objetos"""
    cur = obj
    for part in path.split("."):
        if cur is None: return default
        if isinstance(cur, dict): cur = cur.get(part)
        else: cur = getattr(cur, part, None)
    return default if cur is None else cur

def _iter_disciplinas(instrutor: Any):
    if isinstance(instrutor, dict): return instrutor.get("disciplinas") or []
    return getattr(instrutor, "disciplinas", []) or []

def _apply_border_to_range(ws, cell_range_string, border_style):
    min_col, min_row, max_col, max_row = range_boundaries(cell_range_string)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border_style

# --- Função Principal Atualizada ---
def gerar_mapa_gratificacao_xlsx(
    dados: Iterable[Any], valor_hora_aula: float, nome_mes_ano: str, titulo_curso: str,
    opm_nome: str, escola_nome: str, data_emissao: Optional[date],
    telefone: Optional[str] = None, auxiliar_nome: Optional[str] = None,
    comandante_nome: Optional[str] = None, digitador_nome: Optional[str] = None,
    auxiliar_funcao: Optional[str] = None, comandante_funcao: Optional[str] = None,
    *, data_fim: Optional[date] = None, cidade_assinatura: Optional[str] = "Santa Maria",
) -> bytes:
    valor_hora_aula = float(valor_hora_aula or 0)
    wb = Workbook()
    ws = wb.active
    ws.title = "Mapa de Horas"

    # 1. Configurações de Página e Colunas
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.5, bottom=0.5)
    ws.sheet_view.showGridLines = False
    
    ws.column_dimensions['A'].width = 18 # Posto
    ws.column_dimensions['B'].width = 12 # Id Func
    ws.column_dimensions['C'].width = 35 # Nome
    ws.column_dimensions['D'].width = 30 # Disciplina
    ws.column_dimensions['E'].width = 10 # CH Total
    ws.column_dimensions['F'].width = 15 # CH Anterior
    ws.column_dimensions['G'].width = 10 # CH Pagar
    ws.column_dimensions['H'].width = 15 # Valor R$

    # 2. Definição de Estilos
    thin_border_side = Side(style="thin", color="000000")
    border_all = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    center_align_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_top_align_wrap = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left_align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align_wrap = Alignment(horizontal="right", vertical="center", wrap_text=True)
    left_top_align_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    header_font = Font(name='Times New Roman', size=11)
    table_header_font = Font(name='Calibri', bold=True, size=11)
    table_data_font = Font(name='Calibri', size=11)
    total_font = Font(name='Calibri', bold=True, size=11)
    signature_font = Font(name='Times New Roman', size=11)

    # 3. Construção do Cabeçalho
    ws.merge_cells("A1:B8")
    ws.cell(1, 1).value = f"\n\n\n\n\n________________________\n{comandante_nome or 'Nome Comandante'}\n{comandante_funcao or 'Comandante da EsFAS'}"
    ws.cell(1, 1).font = Font(name='Times New Roman', bold=True, size=10)
    ws.cell(1, 1).alignment = center_top_align_wrap
    _apply_border_to_range(ws, "A1:B8", border_all)

    ws.merge_cells("C1:D5")
    header_text = f"MAPA DE GRATIFICAÇÃO MAGISTÉRIO\nOPM: {opm_nome}\nTelefone: {telefone or 'N/D'}\nHoras aulas a pagar do Mês de {nome_mes_ano}\n{titulo_curso}"
    ws.cell(1, 3).value = header_text
    ws.cell(1, 3).font = header_font
    ws.cell(1, 3).alignment = center_align_wrap

    ws.merge_cells("E1:H8")
    rhe_text = "LANÇAR NO RHE\n____/_____/_____\n\n\n\n\n____________________\nCh da SEÇÃO ADM DE"
    ws.cell(1, 5).value = rhe_text
    ws.cell(1, 5).font = Font(name='Times New Roman', bold=True, size=11)
    ws.cell(1, 5).alignment = center_top_align_wrap
    _apply_border_to_range(ws, "E1:H8", border_all)
    
    current_row = 9

    # 4. Cabeçalho da Tabela
    headers = ["Posto / graduação", "Id. Func.", "Nome completo do servidor", "Disciplina", "CH total", "CH paga anteriormente", "CH a pagar", "Valor em R$"]
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(current_row, col_idx, text)
        cell.font = table_header_font
        cell.alignment = center_align_wrap
        cell.border = border_all
    current_row += 1

    # 5. Dados da Tabela
    total_ch_a_pagar_sum = 0.0
    total_valor_sum = 0.0
    
    if dados:
        # Ordenação por ID Funcional (Matrícula)
        for instrutor in sorted(dados, key=lambda i: str(_safe(i, "matricula", ""))):
            for disc in _iter_disciplinas(instrutor):
                ws.row_dimensions[current_row].height = 25
                
                # Coleta valores (podem vir da edição manual)
                ch_total = float(_safe(disc, "ch_total_disciplina", 0))
                ch_anterior = float(_safe(disc, "ch_anterior", 0))
                ch_mes = float(_safe(disc, "ch_mes", 0))
                valor_a_pagar = ch_mes * valor_hora_aula

                row_data = [
                    _safe(instrutor, "posto", "N/D"),
                    _safe(instrutor, "matricula", ""),
                    _safe(instrutor, "nome", ""),
                    _safe(disc, "nome_disciplina", ""),
                    ch_total,
                    ch_anterior,
                    ch_mes, # CH a pagar este mês
                    valor_a_pagar
                ]

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=value)
                    cell.font = table_data_font
                    cell.border = border_all
                    if col_idx in [1, 2, 5, 6, 7]: cell.alignment = center_align_wrap
                    else: cell.alignment = left_align_wrap
                    
                    if col_idx == 8: cell.number_format = 'R$ #,##0.00'
                    elif col_idx in [5, 6, 7]: cell.number_format = '0.0'

                total_ch_a_pagar_sum += ch_mes
                total_valor_sum += valor_a_pagar
                current_row += 1
    else:
        ws.merge_cells(f"A{current_row}:H{current_row}")
        cell = ws.cell(current_row, 1, "Nenhum dado encontrado.")
        cell.alignment = center_align_wrap; cell.border = border_all
        current_row += 1

    # 6. Linha de Totais
    ws.merge_cells(f"A{current_row}:F{current_row}")
    cell = ws.cell(current_row, 1, "CARGA HORARIA TOTAL")
    cell.font = total_font; cell.alignment = right_align_wrap; cell.border = border_all
    
    ws.cell(current_row, 7, total_ch_a_pagar_sum).font = total_font
    ws.cell(current_row, 7).alignment = center_align_wrap; ws.cell(current_row, 7).border = border_all
    
    ws.cell(current_row, 8, total_valor_sum).font = total_font
    ws.cell(current_row, 8).alignment = right_align_wrap; ws.cell(current_row, 8).border = border_all; ws.cell(current_row, 8).number_format = 'R$ #,##0.00'
    current_row += 2

    # 7. Rodapé de Assinaturas
    bottom_block_start_row = current_row
    ws.merge_cells(f"A{bottom_block_start_row}:F{bottom_block_start_row + 8}")
    orientacoes_text = "ORIENTAÇÕES:\n1. Id. Func. em ordem crescente.\n2. Mapa deverá dar entrada no DE até o dia 05 de cada mês."
    cell = ws.cell(bottom_block_start_row, 1, orientacoes_text)
    cell.alignment = left_top_align_wrap; cell.border = border_all
    _apply_border_to_range(ws, f"A{bottom_block_start_row}:F{bottom_block_start_row + 8}", border_all)

    ws.merge_cells(f"G{bottom_block_start_row}:H{bottom_block_start_row + 8}")
    data_str = data_fim.strftime('%d de %B de %Y') if data_fim else "____/____/____"
    assinaturas_text = f"Quartel em {cidade_assinatura}, {data_str}.\n\n\n____________________\n{auxiliar_nome or 'Auxiliar'}\nDigitador: {digitador_nome or ''}"
    cell = ws.cell(bottom_block_start_row, 7, assinaturas_text)
    cell.alignment = center_top_align_wrap; cell.border = border_all
    _apply_border_to_range(ws, f"G{bottom_block_start_row}:H{bottom_block_start_row + 8}", border_all)

    # 8. Exportação
    out = BytesIO()
    wb.save(out)
    return out.getvalue()

# --- Nova Função: Gerar Quadro de Horário ---
def gerar_quadro_horario_xlsx(pelotao, semana, horario_matrix, datas_semana, tempos, intervalos):
    """
    Gera um arquivo Excel (.xlsx) em memória baseado na matriz do quadro de horários.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Horário - {pelotao}"
    
    # Exibir as linhas de grade padrão
    ws.sheet_view.showGridLines = True
    
    # --- ESTILOS ---
    font_titulo = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    font_header = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    font_corpo = Font(name='Arial', size=10)
    font_intervalo = Font(name='Arial', size=10, italic=True, bold=True, color='555555')
    font_tempo = Font(name='Arial', size=9, bold=True)
    
    fill_titulo = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid') # Azul
    fill_header = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid') # Grafite
    fill_intervalo = PatternFill(start_color='F2F4F4', end_color='F2F4F4', fill_type='solid') # Cinza
    
    border_fina = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7')
    )
    
    # --- CABEÇALHO PRINCIPAL ---
    # Conta quantas colunas teremos (Tempo + 5 dias úteis + fds se houver)
    total_colunas = 6
    if getattr(semana, 'mostrar_sabado', False): total_colunas += 1
    if getattr(semana, 'mostrar_domingo', False): total_colunas += 1
    
    letra_ultima_coluna = get_column_letter(total_colunas)
    
    ws.merge_cells(f'A1:{letra_ultima_coluna}1')
    data_formatada = semana.data_inicio.strftime('%d/%m/%Y') if semana.data_inicio else 'N/D'
    ws['A1'] = f"QUADRO DE HORÁRIOS - PELOTÃO: {pelotao} (Semana de {data_formatada})"
    ws['A1'].font = font_titulo
    ws['A1'].fill = fill_titulo
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    # --- COLUNAS DOS DIAS DA SEMANA ---
    colunas_textos = ['Tempo/Horário', 
                      f"Segunda\n({datas_semana.get('segunda', '')})",
                      f"Terça\n({datas_semana.get('terca', '')})",
                      f"Quarta\n({datas_semana.get('quarta', '')})",
                      f"Quinta\n({datas_semana.get('quinta', '')})",
                      f"Sexta\n({datas_semana.get('sexta', '')})"]
                      
    if getattr(semana, 'mostrar_sabado', False):
        colunas_textos.append(f"Sábado\n({datas_semana.get('sabado', '')})")
    if getattr(semana, 'mostrar_domingo', False):
        colunas_textos.append(f"Domingo\n({datas_semana.get('domingo', '')})")
        
    for col_idx, texto in enumerate(colunas_textos, 1):
        cell = ws.cell(row=2, column=col_idx, value=texto)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border_fina
    ws.row_dimensions[2].height = 35
    
    # --- PREENCHIMENTO DOS DADOS ---
    row_atual = 3
    # Mapeamento dos dias em matriz para a planilha (0=Segunda, 1=Terça...)
    dias_loop = list(range(total_colunas - 1))
    
    # Posições de intervalo convertidas para zero-index (igual ao HTML)
    pos_int_1 = int(intervalos.get('pos_int_1', 3)) - 1
    pos_almoco = int(intervalos.get('pos_almoco', 6)) - 1
    pos_int_2 = int(intervalos.get('pos_int_2', 9)) - 1

    for row_idx in range(15): # 15 períodos possíveis
        periodo_num = row_idx + 1
        
        # Lógica para ocultar períodos noturnos se não estiverem ativos na semana
        if periodo_num > 12:
            if periodo_num == 13 and not getattr(semana, 'mostrar_periodo_13', False): continue
            if periodo_num == 14 and not getattr(semana, 'mostrar_periodo_14', False): continue
            if periodo_num == 15 and not getattr(semana, 'mostrar_periodo_15', False): continue

        # 1. Célula de Tempo (Coluna A)
        tempo_str = f"{tempos[row_idx][0]}\n{tempos[row_idx][1]}"
        cell_tempo = ws.cell(row=row_atual, column=1, value=tempo_str)
        cell_tempo.font = font_tempo
        cell_tempo.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell_tempo.border = border_fina
        
        # 2. Células das Aulas (Colunas B em diante)
        for d_idx, col_matriz in enumerate(dias_loop):
            col_planilha = d_idx + 2
            
            # Verifica se passa do limite de períodos do FDS
            pular_celula = False
            if col_matriz == 5 and periodo_num > getattr(semana, 'periodos_sabado', 0): pular_celula = True
            if col_matriz == 6 and periodo_num > getattr(semana, 'periodos_domingo', 0): pular_celula = True
            
            cell_aula = ws.cell(row=row_atual, column=col_planilha)
            cell_aula.border = border_fina
            cell_aula.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if not pular_celula:
                try:
                    aula = horario_matrix[row_idx][col_matriz]
                except (IndexError, KeyError):
                    aula = None
                
                if aula and aula != 'SKIP':
                    # Extrai os dados reais se a aula existir
                    disciplina = getattr(aula, 'disciplina', {}).get('materia', 'N/D') if isinstance(getattr(aula, 'disciplina', None), dict) else getattr(getattr(aula, 'disciplina', None), 'materia', 'N/D')
                    if disciplina == 'N/D' and isinstance(aula, dict):
                        disciplina = aula.get('disciplina_nome', 'N/D')
                    
                    # Nome do instrutor
                    instrutor = ""
                    if isinstance(aula, dict) and 'instrutores_nomes' in aula:
                        instrutor = "\n".join(aula['instrutores_nomes'])
                    elif getattr(aula, 'instrutor_1', None):
                        instrutor = getattr(aula.instrutor_1.user, 'nome_de_guerra', 'Instrutor')
                        if getattr(aula, 'instrutor_2', None):
                            instrutor += f" / {getattr(aula.instrutor_2.user, 'nome_de_guerra', 'Instrutor')}"
                            
                    cell_aula.value = f"{disciplina}\n{instrutor}"
                    cell_aula.font = font_corpo
                elif aula == 'SKIP':
                    cell_aula.value = "↳" # Indica que é continuação da aula acima
                    cell_aula.font = Font(color="CCCCCC")
            else:
                cell_aula.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

        ws.row_dimensions[row_atual].height = 40
        row_atual += 1
        
        # 3. Inserção de Linhas de Intervalo
        if row_idx == pos_int_1 and intervalos.get('intervalo_1') and intervalos.get('intervalo_1') not in ['N/D', '-', '']:
            ws.merge_cells(start_row=row_atual, start_column=1, end_row=row_atual, end_column=total_colunas)
            c = ws.cell(row=row_atual, column=1, value=f"INTERVALO MANHÃ: {intervalos.get('intervalo_1')}")
            c.font, c.fill, c.alignment = font_intervalo, fill_intervalo, Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row_atual].height = 20
            row_atual += 1
            
        if row_idx == pos_almoco and intervalos.get('almoco') and intervalos.get('almoco') not in ['N/D', '-', '']:
            ws.merge_cells(start_row=row_atual, start_column=1, end_row=row_atual, end_column=total_colunas)
            c = ws.cell(row=row_atual, column=1, value=f"ALMOÇO: {intervalos.get('almoco')}")
            c.font, c.fill, c.alignment = font_intervalo, fill_intervalo, Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row_atual].height = 20
            row_atual += 1
            
        if row_idx == pos_int_2 and intervalos.get('intervalo_2') and intervalos.get('intervalo_2') not in ['N/D', '-', '']:
            ws.merge_cells(start_row=row_atual, start_column=1, end_row=row_atual, end_column=total_colunas)
            c = ws.cell(row=row_atual, column=1, value=f"INTERVALO TARDE: {intervalos.get('intervalo_2')}")
            c.font, c.fill, c.alignment = font_intervalo, fill_intervalo, Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[row_atual].height = 20
            row_atual += 1

    # --- AJUSTE DE LARGURA DAS COLUNAS ---
    ws.column_dimensions['A'].width = 14
    for col_letter in [get_column_letter(i) for i in range(2, total_colunas + 1)]:
        ws.column_dimensions[col_letter].width = 22

    # --- RETORNO DOS BYTES ---
    out = BytesIO()
    wb.save(out)
    return out.getvalue()
